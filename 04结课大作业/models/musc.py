import os
import sys
import numpy as np
import torch
import torch.nn.functional as F
from torch import nn
sys.path.append('./models/backbone')

import datasets.mvtec as mvtec
from datasets.mvtec import _CLASSNAMES as _CLASSNAMES_mvtec_ad
import datasets.visa as visa
from datasets.visa import _CLASSNAMES as _CLASSNAMES_visa
import datasets.btad as btad
from datasets.btad import _CLASSNAMES as _CLASSNAMES_btad
from models.backbone.EfficientVit.classification.model.build import EfficientViT_M4
from models.backbone.EfficientFormer.models.efficientformer_v2 import efficientformerv2_s1
import models.backbone.open_clip as open_clip
import models.backbone._backbones as _backbones
from models.modules._LNAMD import LNAMD
from models.modules._MSM import MSM
from models.modules._RsCIN import RsCIN
from utils.metrics import compute_metrics
from openpyxl import Workbook
from tqdm import tqdm
import pickle
import copy
import time
import cv2

import logging
from datetime import datetime

import warnings
warnings.filterwarnings("ignore")





class MuSc():
    def __init__(self, cfg, seed=0):
        self.cfg = cfg
        self.seed = seed
        if cfg['device'] =='cpu':
            self.device = torch.device("cpu")
        else:
            self.device = torch.device("cuda:{}".format(cfg['device']) if torch.cuda.is_available() else "cpu")


        self.path = cfg['datasets']['data_path']
        self.dataset = cfg['datasets']['dataset_name']
        self.vis = cfg['testing']['vis']
        self.vis_type = cfg['testing']['vis_type']
        self.save_excel = cfg['testing']['save_excel']
        # the categories to be tested
        self.categories = cfg['datasets']['class_name']
        if isinstance(self.categories, str):
            if self.categories.lower() == 'all':
                if self.dataset == 'visa':
                    self.categories = _CLASSNAMES_visa
                elif self.dataset == 'mvtec_ad':
                    self.categories = _CLASSNAMES_mvtec_ad
                elif self.dataset == 'btad':
                    self.categories = _CLASSNAMES_btad
            else:
                self.categories = [self.categories]

        self.model_name = cfg['models']['backbone_name']
        self.image_size = cfg['datasets']['img_resize']
        self.batch_size = cfg['models']['batch_size']
        self.pretrained = cfg['models']['pretrained']
        self.features_list = [l+1 for l in cfg['models']['feature_layers']]
        self.divide_num = cfg['datasets']['divide_num']
        self.r_list = cfg['models']['r_list']
        self.output_dir = os.path.join(cfg['testing']['output_dir'], self.dataset, self.model_name, 'imagesize{}'.format(self.image_size))
        os.makedirs(self.output_dir, exist_ok=True)

        self._setup_logger(cfg)
        
        self.load_backbone()
        self.total_lnamd_time = 0.0
        self.total_msm_time = 0.0
        if any(x in self.model_name for x in ['efficientvit', 'efficientformer', 'mobilevit','fastvit']):
            self.is_cnn_backbone = True
        else:
            self.is_cnn_backbone = False
    def _setup_logger(self, cfg):
        log_dir = os.path.join('logs', self.model_name, self.dataset)
        os.makedirs(log_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        log_filename = f"imgresize_{self.image_size}_{timestamp}.log"
        log_path = os.path.join(log_dir, log_filename)
        self.logger = logging.getLogger(f"MuSc.{self.model_name}.{timestamp}")
        self.logger.setLevel(logging.INFO)
        if self.logger.hasHandlers():
            self.logger.handlers.clear()
        formatter = logging.Formatter('%(message)s')

        fh = logging.FileHandler(log_path, mode='a')
        fh.setLevel(logging.INFO)
        fh.setFormatter(formatter)
        self.logger.addHandler(fh)

        sh = logging.StreamHandler(sys.stdout)
        sh.setLevel(logging.INFO)
        sh.setFormatter(formatter)
        self.logger.addHandler(sh)
        # log configuration
        ds = cfg.get('datasets', {})
        ms = cfg.get('models', {})
        ts = cfg.get('testing', {})
        self.logger.info("===== Configuration =====")
        self.logger.info(f"dataset_name: {ds.get('dataset_name')}")
        self.logger.info(f"data_path: {ds.get('data_path')}")
        self.logger.info(f"class_name: {ds.get('class_name')}")
        self.logger.info(f"img_resize: {ds.get('img_resize')}")
        self.logger.info(f"divide_num: {ds.get('divide_num')}")
        self.logger.info(f"backbone_name: {ms.get('backbone_name')}")
        self.logger.info(f"pretrained: {ms.get('pretrained')}")
        self.logger.info(f"batch_size: {ms.get('batch_size')}")
        self.logger.info(f"feature_layers: {ms.get('feature_layers')}")
        self.logger.info(f"r_list: {ms.get('r_list')}")
        self.logger.info(f"device: {cfg.get('device')}")
        self.logger.info(f"output_dir: {ts.get('output_dir')}")
        self.logger.info(f"vis: {ts.get('vis')}")
        self.logger.info(f"vis_type: {ts.get('vis_type')}")
        self.logger.info(f"save_excel: {ts.get('save_excel')}")
        self.logger.info("===========backbone_info=============")

    def _print_model_params(self, model, name: str):
        """Print total/trainable parameter count in millions (M). Safe for models without parameters."""
        if model is None:
            self.logger.info(f"[{name}] params(M): total=0.00M trainable=0.00M")
            return
        try:
            params = list(model.parameters())
        except Exception:
            self.logger.info(f"[{name}] params(M): <unable to enumerate parameters>")
            return
        total_params = sum(p.numel() for p in params)
        trainable_params = sum(p.numel() for p in params if getattr(p, "requires_grad", False))
        self.logger.info(f"[{name}] params(M): total={total_params/1e6:.2f}M trainable={trainable_params/1e6:.2f}M")

    def _print_feature_params(self, model, name: str, last_layer: str):
        """Print parameter count up to and including last_layer (top-level module name or 'block.idx')."""
        top_key = last_layer.split(".")[0]
        total = 0
        for key, module in model._modules.items():
            total += sum(p.numel() for p in module.parameters())
            if key == top_key:
                break
        self.logger.info(f"[{name}] params_only_feature(M): {total/1e6:.2f}M (up to '{top_key}')")

    def _calculate_vit_flops(self, visual, name: str, image_size: int):
        """Calculate ViT FLOPs analytically from model attributes.

        Covers: patch embedding (Conv2d), transformer (QKV, attention, out-proj,
        MLP), LN, class token, and final projection. All counted as MACs.
        """
        patch_size = visual.conv1.kernel_size[0]
        width = visual.conv1.out_channels  # transformer hidden dim
        grid = image_size // patch_size
        N = grid * grid + 1  # num_patches + cls token
        layers = len(visual.transformer.resblocks)

        # Infer MLP hidden dim and heads from first block
        blk = visual.transformer.resblocks[0]
        mlp_dim = blk.mlp[0].out_features if hasattr(blk.mlp, '__getitem__') else blk.mlp.c_fc.out_features
        heads = blk.attn.num_heads

        # 1. Patch embedding conv: in_ch * k*k * width * grid * grid
        patch_embed = 3 * patch_size * patch_size * width * grid * grid

        # 2. Per transformer layer
        qkv = 3 * N * width * width                    # Q, K, V projections
        attn = 2 * heads * N * N * (width // heads)     # q@k + attn@v
        out_proj = N * width * width                     # output projection
        mlp_flops = 2 * N * width * mlp_dim              # fc1 + fc2
        ln = 2 * N * width                               # 2 LayerNorms per layer (small)
        per_layer = qkv + attn + out_proj + mlp_flops + ln
        transformer_total = per_layer * layers

        # 3. Final LN + projection
        final_ln = N * width
        proj_dim = visual.proj.shape[1] if visual.proj is not None else 0
        final_proj = width * proj_dim

        total = patch_embed + transformer_total + final_ln + final_proj
        self.logger.info(f"[{name}] FLOPs_feature(G): {total/1e9:.2f}G "
              f"(ViT: {layers}L, d={width}, mlp={mlp_dim}, heads={heads}, "
              f"patches={grid}x{grid}, img={image_size})")

    def _calculate_feature_flops(self, model, name: str, last_layer: str = None, input_shape=(3, 512, 512), **fwd_kwargs):
        """Calculate FLOPs up to and including last_layer for feature extraction.

        If last_layer is None, calculates FLOPs for the entire model.
        Tries fvcore first; falls back to hook-based counting for models with
        custom CUDA/Triton ops or non-standard forward signatures.
        """
        device = next(model.parameters()).device
        label = f"up to '{last_layer}'" if last_layer else "full model"

        # --- When last_layer is specified, try fvcore with a truncated wrapper ---
        if last_layer is not None:
            parts = last_layer.split(".")

            class FeatureExtractionWrapper(torch.nn.Module):
                def __init__(self, backbone, layer_parts):
                    super().__init__()
                    self.backbone = backbone
                    self.layer_parts = layer_parts

                def forward(self, x):
                    if len(self.layer_parts) == 1:
                        for key, module in self.backbone._modules.items():
                            x = module(x)
                            if key == self.layer_parts[0]:
                                break
                    else:
                        parent_key = self.layer_parts[0]
                        sub_key = self.layer_parts[1]
                        for key, module in self.backbone._modules.items():
                            if key == parent_key:
                                for sub_name, sub_module in module._modules.items():
                                    x = sub_module(x)
                                    if sub_name == sub_key:
                                        break
                                break
                            else:
                                x = module(x)
                    return x

            try:
                from fvcore.nn import FlopCountAnalysis
                wrapper = FeatureExtractionWrapper(model, parts)
                wrapper.eval()
                dummy_input = torch.randn(1, *input_shape, device=device)
                flops = FlopCountAnalysis(wrapper, dummy_input)
                flops.unsupported_ops_warnings(False)
                flops.uncalled_modules_warnings(False)
                total = flops.total()
                self.logger.info(f"[{name}] FLOPs_feature(G): {total/1e9:.2f}G ({label})")
                return
            except Exception:
                pass

        # --- fvcore on full model (no wrapper) ---
        try:
            from fvcore.nn import FlopCountAnalysis
            model.eval()
            dummy_input = torch.randn(1, *input_shape, device=device)
            if fwd_kwargs:
                # Wrap model so fvcore sees a standard forward(x) signature
                _model = model
                _kwargs = fwd_kwargs
                class _Wrapper(torch.nn.Module):
                    def __init__(self):
                        super().__init__()
                        self.inner = _model
                    def forward(self, x):
                        return self.inner(x, **_kwargs)
                target = _Wrapper().eval()
            else:
                target = model
            flops = FlopCountAnalysis(target, dummy_input)
            flops.unsupported_ops_warnings(False)
            flops.uncalled_modules_warnings(False)
            total = flops.total()
            self.logger.info(f"[{name}] FLOPs_feature(G): {total/1e9:.2f}G ({label})")
            return
        except Exception:
            pass

        # --- Fallback: hook-based counting on the original model ---
        try:
            total_flops = [0]
            hooks = []

            def hook_fn(module, inp, output):
                if isinstance(module, torch.nn.Conv2d):
                    if isinstance(output, torch.Tensor) and output.dim() == 4:
                        _, _, h_out, w_out = output.shape
                    else:
                        return
                    total_flops[0] += (module.in_channels // module.groups) * module.kernel_size[0] * module.kernel_size[1] * module.out_channels * h_out * w_out
                elif isinstance(module, torch.nn.Linear):
                    total_flops[0] += module.in_features * module.out_features
                elif isinstance(module, (torch.nn.BatchNorm2d, torch.nn.LayerNorm, torch.nn.GroupNorm)):
                    if isinstance(output, torch.Tensor):
                        total_flops[0] += output.numel()

            for m in model.modules():
                if isinstance(m, (torch.nn.Conv2d, torch.nn.Linear, torch.nn.BatchNorm2d, torch.nn.LayerNorm, torch.nn.GroupNorm)):
                    hooks.append(m.register_forward_hook(hook_fn))

            dummy_input = torch.randn(1, *input_shape, device=device)
            with torch.no_grad():
                model(dummy_input, **fwd_kwargs)

            for h in hooks:
                h.remove()

            self.logger.info(f"[{name}] FLOPs_feature(G): {total_flops[0]/1e9:.2f}G ({label}) [hook-based, approx]")
        except Exception as e:
            self.logger.info(f"[{name}] FLOPs calculation failed: {e}")


    def load_backbone(self):
        # DINO / DINOv2 backbones
        if 'dino' in self.model_name:
            self.dino_model = _backbones.load(self.model_name)
            self._print_model_params(self.dino_model, self.model_name)
            self.dino_model.to(self.device)
            self.preprocess = None
        elif 'efficientvit' in self.model_name:
            self.logger.info('efficientvit_m4')
            ckpt = '/home/jwhuang/MuSc/models/backbone/EfficientVit/checkpoints/efficientvit_m4.pth'
            self.model = EfficientViT_M4(pretrained=ckpt)
            self._print_model_params(self.model, self.model_name)
            self.logger.info(self.model._modules.keys())
            self._print_feature_params(self.model, self.model_name, 'blocks3')
            self.model.to(self.device)
            self.model.eval()
            self._calculate_feature_flops(self.model, self.model_name, 'blocks3')
            self.preprocess = None
            self.cnn_backbones_most_dim = 384
            self.cnn_backbones_most_side = 32
            self.extractor = FeatureExtractor(
                                backbone=self.model,
                                layers_to_extract_from=['blocks1'],
                                device=self.device,
                                input_shape=(3, 512, 512),
                                cnn_backbones_most_dim=self.cnn_backbones_most_dim,
                                cnn_backbones_most_side=self.cnn_backbones_most_side)
            self.logger.info("===== FeatureExtractor layers =====")
            self.logger.info(f"FeatureExtractor layers: {self.extractor.layers_to_extract_from}")
            self.features_list = list(range(1, len(self.extractor.layers_to_extract_from) + 1))
            self.logger.info(f"Updated features_list to: {self.features_list}")
            # features.shape torch.Size([4, 128, 32, 32])
            # features.shape torch.Size([4, 256, 16, 16])
            # features.shape torch.Size([4, 384, 8, 8])
        elif 'fastvit_sa12' in self.model_name:
            self.logger.info('fastvit_sa12')
            import timm
            ckpt_path = '/home/jwhuang/MuSc/models/backbone/fastvit/checkpoints/pytorch_model.bin'
            self.model = timm.create_model('fastvit_sa12', pretrained=True, features_only=True,
                                           pretrained_cfg_overlay=dict(file=ckpt_path))
            self._print_model_params(self.model, self.model_name)
            self.logger.info(self.model._modules.keys())
            self._print_feature_params(self.model, self.model_name, 'stages_3')
            self.model.to(self.device)
            self.model.eval()
            self._calculate_feature_flops(self.model, self.model_name, 'stages_3')
            self.preprocess = None
            # 512 input: stage0鈫?28x128(64ch), stage1鈫?4x64(128ch), stage2鈫?2x32(256ch), stage3鈫?6x16(512ch)
            self.cnn_backbones_most_dim = 512
            self.cnn_backbones_most_side = 64
            self.extractor = FeatureExtractor(
                backbone=self.model,
                layers_to_extract_from=['stages_1'],
                device=self.device,
                input_shape=(3, 512, 512),
                cnn_backbones_most_dim=self.cnn_backbones_most_dim,
                cnn_backbones_most_side=self.cnn_backbones_most_side,
            )
            self.logger.info("===== FeatureExtractor layers =====")
            self.logger.info(f"FeatureExtractor layers: {self.extractor.layers_to_extract_from}")
            self.features_list = list(range(1, len(self.extractor.layers_to_extract_from) + 1))
            self.logger.info(f"Updated features_list to: {self.features_list}")
            # patch_features.shape 0 torch.Size([B, 64, 128, 128])
            # patch_features.shape 1 torch.Size([B, 128, 64, 64])
            # patch_features.shape 2 torch.Size([B, 256, 32, 32])
            # patch_features.shape 3 torch.Size([B, 512, 16, 16])
        elif 'swin_tiny' in self.model_name:
            self.logger.info('swin_tiny_patch4_window7_224')
            import timm
            ckpt_path = '/home/jwhuang/MuSc/models/backbone/swintransformer/checkpoints/pytorch_model.bin'
            # img_size=512 required; output is [B,H,W,C] channel-last
            self.model = timm.create_model('swin_tiny_patch4_window7_224', pretrained=True, features_only=True,
                                           img_size=self.image_size, pretrained_cfg_overlay=dict(file=ckpt_path))
            self._print_model_params(self.model, self.model_name)
            self.logger.info(self.model._modules.keys())
            self._print_feature_params(self.model, self.model_name, 'layers_3')
            self.model.to(self.device)
            self.model.eval()
            self._calculate_feature_flops(self.model, self.model_name, 'layers_3')
            self.preprocess = None
            #鎵€鐢ㄧ殑鐗瑰緛灞?            self.swin_feature_indices = [1]
            self.logger.info(f"swin_tiny feature indices: {self.swin_feature_indices}")
            self.features_list = list(range(1, len(self.swin_feature_indices) + 1))
            self.logger.info(f"Updated features_list to: {self.features_list}")
            # 512 input: layers_1鈫抂B,64,64,192], layers_2鈫抂B,32,32,384], layers_3鈫抂B,16,16,768]
            #layer_0 to layer_3
            # patch_features.shape 0 torch.Size([4, 128, 128, 96])
            # patch_features.shape 1 torch.Size([4, 64, 64, 192])
            # patch_features.shape 2 torch.Size([4, 32, 32, 384])
            # patch_features.shape 3 torch.Size([4, 16, 16, 768])
        elif 'mobilevit_s' in self.model_name:
            self.logger.info('mobilevit_s')
            import timm
            ckpt_path = '/home/jwhuang/MuSc/models/backbone/mobilevit/checkpoints/pytorch_model.bin'
            self.model = timm.create_model('mobilevit_s', pretrained=True, features_only=True,
                                           pretrained_cfg_overlay=dict(file=ckpt_path))
            self._print_model_params(self.model, self.model_name)
            self.logger.info(self.model._modules.keys())
            self._print_feature_params(self.model, self.model_name, 'stages_4')
            self.model.to(self.device)
            self.model.eval()
            self._calculate_feature_flops(self.model, self.model_name, 'stages_4')
            self.preprocess = None
            # 512 input: stages_2鈫?4x64(96ch,MobileVitBlock), stages_3鈫?2x32(128ch), final_conv鈫?6x16(640ch)
            # stages_1 is pure CNN (no attention); use stages_2+ which have MobileVitBlock
            self.cnn_backbones_most_dim = 160
            self.cnn_backbones_most_side = 64
            self.extractor = FeatureExtractor(
                backbone=self.model,
                layers_to_extract_from=['stages_2'],
                device=self.device,
                input_shape=(3, 512, 512),
                cnn_backbones_most_dim=self.cnn_backbones_most_dim,
                cnn_backbones_most_side=self.cnn_backbones_most_side,
            )
            self.logger.info("===== FeatureExtractor layers =====")
            self.logger.info(f"FeatureExtractor layers: {self.extractor.layers_to_extract_from}")
            self.features_list = list(range(1, len(self.extractor.layers_to_extract_from) + 1))
            self.logger.info(f"Updated features_list to: {self.features_list}")
            #stage_0 to stage_4 final_conv
            # patch_features.shape 0 torch.Size([3, 32, 256, 256])
            # patch_features.shape 1 torch.Size([3, 64, 128, 128])
            # patch_features.shape 2 torch.Size([3, 96, 64, 64])
            # patch_features.shape 3 torch.Size([3, 128, 32, 32])
            # patch_features.shape 4 torch.Size([3, 160, 16, 16])
            # patch_features.shape 5 torch.Size([3, 640, 16, 16])
        elif 'efficientformerv2_s1' in self.model_name:
            self.logger.info('efficientformerv2_s1')
            ckpt_path = '/home/jwhuang/MuSc/models/backbone/EfficientFormer/checkpoints/eformer_s1_450.pth'
            # Initialize with resolution=512 so Attention4D bias shapes match the actual input
            self.model = efficientformerv2_s1(pretrained=False, resolution=512)
            checkpoint = torch.load(ckpt_path, map_location='cpu')
            state_dict = checkpoint['model'] if 'model' in checkpoint else checkpoint
            model_state = self.model.state_dict()
            state_dict = {k: v for k, v in state_dict.items()
                          if k not in model_state or v.shape == model_state[k].shape}
            self.model.load_state_dict(state_dict, strict=False)
            self._print_model_params(self.model, self.model_name)
            self.logger.info(self.model._modules.keys())
            self._print_feature_params(self.model, self.model_name, 'network.6')
            if hasattr(self.model, "network"):
                try:
                    self.logger.info(f" efficient former network len = {len(self.model.network)}")
                except Exception:
                    pass
            self.model.to(self.device)
            self.model.eval()
            self._calculate_feature_flops(self.model, self.model_name, 'network.6')
            self.preprocess = None
            # EfficientFormerV2-S1 embed_dims=[32,48,120,224]
            # 512 input: network.2鈫?4x64(48ch), network.4鈫?2x32(120ch), network.6鈫?6x16(224ch)
            self.cnn_backbones_most_dim = 224
            self.cnn_backbones_most_side = 16
            self.extractor = FeatureExtractor(
                backbone=self.model,
                layers_to_extract_from=['network.2'],
                device=self.device,
                input_shape=(3, 512, 512),
                cnn_backbones_most_dim=self.cnn_backbones_most_dim,
                cnn_backbones_most_side=self.cnn_backbones_most_side,
            )
            self.logger.info("===== FeatureExtractor layers =====")
            self.logger.info(f"FeatureExtractor layers: {self.extractor.layers_to_extract_from}")
            self.features_list = list(range(1, len(self.extractor.layers_to_extract_from) + 1))
            self.logger.info(f"Updated features_list to: {self.features_list}")
            # patch_features.shape 0 torch.Size([4, 32, 128, 128])
            # patch_features.shape 1 torch.Size([4, 48, 64, 64])
            # patch_features.shape 2 torch.Size([4, 48, 64, 64])
            # patch_features.shape 3 torch.Size([4, 120, 32, 32])
            # patch_features.shape 4 torch.Size([4, 120, 32, 32])
            # patch_features.shape 5 torch.Size([4, 224, 16, 16])
            # patch_features.shape 6 torch.Size([4, 224, 16, 16])
        elif self.model_name in _backbones._BACKBONES:
            self.dino_model = _backbones.load(self.model_name)
            self._print_model_params(self.dino_model, self.model_name)
            self.dino_model.to(self.device)
            self._calculate_feature_flops(self.dino_model, self.model_name)
            self.preprocess = None
        else:
            # CLIP / open_clip backbones (ViT-B/L, etc.)
            self.clip_model, _, self.preprocess = open_clip.create_model_and_transforms(
                self.model_name, self.image_size, pretrained=self.pretrained
            )
            # open_clip model includes both visual/text towers; print total as a quick sanity-check.
            self._print_model_params(self.clip_model, self.model_name)
            if hasattr(self.clip_model, 'visual'):
                self._print_model_params(self.clip_model.visual, f"{self.model_name}_visual_only")
            self.clip_model.to(self.device)
            if hasattr(self.clip_model, 'visual'):
                self._calculate_vit_flops(self.clip_model.visual, f"{self.model_name}_visual_only", self.image_size)
            #[B,1297,1024]#512杈撳叆

    def load_datasets(self, category, divide_num=1, divide_iter=0):
        # dataloader
        if self.dataset == 'visa':
            test_dataset = visa.VisaDataset(source=self.path, split=visa.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        elif self.dataset == 'mvtec_ad':
            test_dataset = mvtec.MVTecDataset(source=self.path, split=mvtec.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        elif self.dataset == 'btad':
            test_dataset = btad.BTADDataset(source=self.path, split=btad.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        return test_dataset


    def visualization(self, image_path_list, gt_list, pr_px, category):
        def normalization01(img):
            return (img - img.min()) / (img.max() - img.min())
        if self.vis_type == 'single_norm':
            # normalized per image
            for i, path in enumerate(image_path_list):
                anomaly_type = path.split('/')[-2]
                img_name = path.split('/')[-1]
                if anomaly_type not in ['good', 'Normal', 'ok'] and gt_list[i] != 0:
                    save_path = os.path.join(self.output_dir, category, anomaly_type)
                    os.makedirs(save_path, exist_ok=True)
                    save_path = os.path.join(save_path, img_name)
                    anomaly_map = pr_px[i].squeeze()
                    anomaly_map = normalization01(anomaly_map)*255
                    anomaly_map = cv2.applyColorMap(anomaly_map.astype(np.uint8), cv2.COLORMAP_JET)
                    cv2.imwrite(save_path, anomaly_map)
        else:
            # normalized all image
            pr_px = normalization01(pr_px)
            for i, path in enumerate(image_path_list):
                anomaly_type = path.split('/')[-2]
                img_name = path.split('/')[-1]
                save_path = os.path.join(self.output_dir, category, anomaly_type)
                os.makedirs(save_path, exist_ok=True)
                save_path = os.path.join(save_path, img_name)
                anomaly_map = pr_px[i].squeeze()
                anomaly_map *= 255
                anomaly_map = cv2.applyColorMap(anomaly_map.astype(np.uint8), cv2.COLORMAP_JET)
                cv2.imwrite(save_path, anomaly_map)


    def make_category_data(self, category):
        self.logger.info(category)

        # divide sub-datasets
        divide_num = self.divide_num
        anomaly_maps = torch.tensor([]).double()
        gt_list = []
        img_masks = []
        class_tokens = []
        image_path_list = []
        start_time_all = time.time()
        dataset_num = 0
        for divide_iter in range(divide_num):
            # Accumulate feature extraction time for this sub-dataset only (will be averaged by subset_num below)
            feature_extract_time_only_backbone = 0.0
            test_dataset = self.load_datasets(category, divide_num=divide_num, divide_iter=divide_iter)
            test_dataloader = torch.utils.data.DataLoader(
                test_dataset,
                batch_size=self.batch_size,
                shuffle=False,
                num_workers=0,
                pin_memory=True,
            )
            
            # extract features
            patch_tokens_list = []
            subset_num = len(test_dataset)
            dataset_num += subset_num

            # -------------------------
            # Warmup (not timed, not saved; keeps first batch for real loop)
            # Configure via cfg['testing']['warmup_steps'] (default: 2). Set 0 to disable.
            # -------------------------
            warmup_steps = int(self.cfg.get('testing', {}).get('warmup_steps', 2))
            data_iter = iter(test_dataloader)
            first_batch = next(data_iter, None)
            if first_batch is not None and warmup_steps > 0:
                def _get_image_tensor(batch):
                    if isinstance(batch, dict):
                        return batch["image"]
                    return batch[0]

                warm_image = _get_image_tensor(first_batch)
                with torch.no_grad(), torch.cuda.amp.autocast():
                    warm_input = warm_image.to(torch.float).to(self.device)
                    for _ in range(warmup_steps):
                        if self.is_cnn_backbone:
                            _ = self.extractor.forward_modules["feature_aggregator"](warm_input)
                        elif 'dinov2' in self.model_name:
                            _ = self.dino_model.get_intermediate_layers(
                                x=warm_input,
                                n=[l - 1 for l in self.features_list],
                                return_class_token=False,
                            )
                            _ = self.dino_model(warm_input)
                        elif 'dino' in self.model_name:
                            _ = self.dino_model.get_intermediate_layers(x=warm_input, n=max(self.features_list))
                            _ = self.dino_model(warm_input)
                        elif 'swin_tiny' in self.model_name:
                            _ = self.model(warm_input)
                        else:  # clip
                            _ = self.clip_model.encode_image(warm_input, self.features_list)
                if torch.cuda.is_available():
                    torch.cuda.synchronize()

            # Iterate including the first batch we already consumed, without dropping data
            if first_batch is None:
                data_stream = []
                total_batches = 0
            else:
                import itertools  # local import to keep warmup localized
                data_stream = itertools.chain([first_batch], data_iter)
                total_batches = len(test_dataloader)
            # ----------------------  warmup end ------------------------------------- 
            # torch.backends.cudnn.benchmark = True
            start_time = time.time()
            # for image_info in tqdm(test_dataloader):
            for image_info in tqdm(data_stream, total=total_batches):#warmup
            # for image_info in test_dataloader:
                if isinstance(image_info, dict):
                    image = image_info["image"]
                    image_path_list.extend(image_info["image_path"])
                    img_masks.append(image_info["mask"])
                    gt_list.extend(list(image_info["is_anomaly"].numpy()))
                with torch.no_grad(), torch.cuda.amp.autocast():
                    input_image = image.to(torch.float).to(self.device)
                    if self.is_cnn_backbone:
                        patch_features, cnn_extract_time = self.extractor.embed_no_aggregation(input_image)
                        feature_extract_time_only_backbone += cnn_extract_time

                        patch_tokens = []
                        for feat in patch_features:
                            B, C, H, W = feat.shape
                            feat_seq = feat.permute(0, 2, 3, 1).reshape(B, H * W, C)
                            fake_cls = torch.zeros(B, 1, C, device=feat.device)
                            tokens = torch.cat([fake_cls, feat_seq], dim=1)
                            patch_tokens.append(tokens.cpu())
                        image_features = torch.mean(patch_tokens[-1][:, 1:, :].to(self.device), dim=1)
                        image_features = image_features / image_features.norm(dim=-1, keepdim=True)
                        image_features = image_features.cpu().numpy()
                        image_features = [image_features[i] for i in range(image_features.shape[0])]
                        class_tokens.extend(image_features)
                        patch_tokens_list.append(patch_tokens)
                    elif 'swin_tiny' in self.model_name:
                        start_time_swin = time.time()
                        feats = self.model(input_image)  # list of [B,H,W,C] per stage
                        feature_extract_time_only_backbone += time.time() - start_time_swin
                        selected = [feats[i].permute(0, 3, 1, 2).contiguous() for i in self.swin_feature_indices]
                        patch_tokens = []
                        for feat in selected:
                            B, C, H, W = feat.shape
                            feat_seq = feat.permute(0, 2, 3, 1).reshape(B, H * W, C)
                            fake_cls = torch.zeros(B, 1, C, device=feat.device)
                            tokens = torch.cat([fake_cls, feat_seq], dim=1)
                            patch_tokens.append(tokens.cpu())
                        image_features = torch.mean(patch_tokens[-1][:, 1:, :].to(self.device), dim=1)
                        image_features = image_features / image_features.norm(dim=-1, keepdim=True)
                        image_features = image_features.cpu().numpy()
                        image_features = [image_features[i] for i in range(image_features.shape[0])]
                        class_tokens.extend(image_features)
                        patch_tokens_list.append(patch_tokens)
                    elif 'dinov2' in self.model_name:
                        patch_tokens = self.dino_model.get_intermediate_layers(x=input_image, n=[l-1 for l in self.features_list], return_class_token=False)
                        image_features = self.dino_model(input_image)
                        patch_tokens = [patch_tokens[l].cpu() for l in range(len(self.features_list))]
                        fake_cls = [torch.zeros_like(p)[:, 0:1, :] for p in patch_tokens]
                        patch_tokens = [torch.cat([fake_cls[i], patch_tokens[i]], dim=1) for i in range(len(patch_tokens))]
                    elif 'dino' in self.model_name:
                        patch_tokens_all = self.dino_model.get_intermediate_layers(x=input_image, n=max(self.features_list))
                        image_features = self.dino_model(input_image)
                        patch_tokens = [patch_tokens_all[l-1].cpu() for l in self.features_list]
                    else: # clip
                        start_clip_time = time.time()
                        image_features, patch_tokens = self.clip_model.encode_image(input_image, self.features_list)
                        feature_extract_time_only_backbone += time.time() - start_clip_time
                        image_features /= image_features.norm(dim=-1, keepdim=True)
                        patch_tokens = [patch_tokens[l].cpu() for l in range(len(self.features_list))]

                if not self.is_cnn_backbone and 'swin_tiny' not in self.model_name:
                    image_features = [image_features[bi].squeeze().cpu().numpy() for bi in range(image_features.shape[0])]
                    class_tokens.extend(image_features)
                    patch_tokens_list.append(patch_tokens)

            end_time = time.time()
            self.logger.info(f'subset_num: {subset_num}')
            self.logger.info('extract time: {}ms per image'.format((end_time-start_time)*1000/subset_num))
            self.logger.info('feature extract only backbone time: {}ms per image'.format(feature_extract_time_only_backbone*1000/subset_num))
            # LNAMD
            feature_dim = patch_tokens_list[0][0].shape[-1]
            anomaly_maps_r = torch.tensor([]).double()
            for r in self.r_list:
                start_time = time.time()
                print('aggregation degree: {}'.format(r))
                LNAMD_r = LNAMD(device=self.device, r=r, feature_dim=feature_dim, feature_layer=self.features_list)
                Z_layers = {}
                for im in range(len(patch_tokens_list)):
                    patch_tokens = [p.to(self.device) for p in patch_tokens_list[im]]
                    with torch.no_grad(), torch.cuda.amp.autocast():
                        features = LNAMD_r._embed(patch_tokens)
                        features /= features.norm(dim=-1, keepdim=True)
                        for l in range(len(self.features_list)):
                            # save the aggregated features
                            if str(l) not in Z_layers.keys():
                                Z_layers[str(l)] = []
                            Z_layers[str(l)].append(features[:, :, l, :])
                end_time = time.time()
                print('LNAMD-{}: {}ms per image'.format(r, (end_time-start_time)*1000/subset_num))
                self.total_lnamd_time += (end_time-start_time)*1000/subset_num

                # MSM
                anomaly_maps_l = torch.tensor([]).double()
                start_time = time.time()
                for l in Z_layers.keys():
                    # different layers
                    Z = torch.cat(Z_layers[l], dim=0).to(self.device) # (N, L, C)
                    print('layer-{} mutual scoring...'.format(l))
                    anomaly_maps_msm = MSM(Z=Z, device=self.device, topmin_min=0, topmin_max=0.3)
                    anomaly_maps_l = torch.cat((anomaly_maps_l, anomaly_maps_msm.unsqueeze(0).cpu()), dim=0)
                    torch.cuda.empty_cache()
                anomaly_maps_l = torch.mean(anomaly_maps_l, 0)
                anomaly_maps_r = torch.cat((anomaly_maps_r, anomaly_maps_l.unsqueeze(0)), dim=0)
                end_time = time.time()
                print('MSM: {}ms per image'.format((end_time-start_time)*1000/subset_num))
                self.total_msm_time += (end_time-start_time)*1000/subset_num
            anomaly_maps_iter = torch.mean(anomaly_maps_r, 0).to(self.device)
            del anomaly_maps_r
            torch.cuda.empty_cache()

            # interpolate
            B, L = anomaly_maps_iter.shape
            H = int(np.sqrt(L))
            anomaly_maps_iter = F.interpolate(anomaly_maps_iter.view(B, 1, H, H),
                                        size=self.image_size, mode='bilinear', align_corners=True)
            anomaly_maps = torch.cat((anomaly_maps, anomaly_maps_iter.cpu()), dim=0)

        # save image features for optimizing classification
        # cls_save_path = os.path.join('./image_features/{}_{}.dat'.format(dataset, category))
        # with open(cls_save_path, 'wb') as f:
        #     pickle.dump([np.array(class_tokens)], f)
        end_time_all = time.time()
        self.logger.info('LNAMD time: {}ms'.format(self.total_lnamd_time))
        self.logger.info('MSM time: {}ms'.format(self.total_msm_time))
        self.logger.info('MuSc: {}ms per image'.format((end_time_all-start_time_all)*1000/dataset_num))

        anomaly_maps = anomaly_maps.cpu().numpy()
        torch.cuda.empty_cache()

        B = anomaly_maps.shape[0]   # the number of unlabeled test images
        ac_score = np.array(anomaly_maps).reshape(B, -1).max(-1)
        # RsCIN
        # if self.dataset == 'visa':
        #     k_score = [1, 8, 9]
        # elif self.dataset == 'mvtec_ad':
        #     k_score = [1, 2, 3]
        # else:
        #     k_score = [1, 2, 3]
        # scores_cls = RsCIN(ac_score, class_tokens, k_list=k_score)

        self.logger.info('computing metrics...')
        pr_sp = np.array(ac_score)
        gt_sp = np.array(gt_list)
        gt_px = torch.cat(img_masks, dim=0).numpy().astype(np.int32)
        pr_px = np.array(anomaly_maps)
        image_metric, pixel_metric = compute_metrics(gt_sp, pr_sp, gt_px, pr_px)
        auroc_sp, f1_sp, ap_sp = image_metric
        auroc_px, f1_px, ap_px, aupro = pixel_metric
        self.logger.info(category)
        self.logger.info('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp*100, f1_sp*100, ap_sp*100))
        self.logger.info('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px*100, f1_px*100, ap_px*100, aupro*100))

        if self.vis:
            self.logger.info('visualization...')
            self.visualization(image_path_list, gt_list, pr_px, category)
    
        return image_metric, pixel_metric


    def main(self):
        auroc_sp_ls = []
        f1_sp_ls = []
        ap_sp_ls = []
        auroc_px_ls = []
        f1_px_ls = []
        ap_px_ls = []
        aupro_ls = []
        self.logger.info("===========start_MuSc=============")
        for category in self.categories:
            image_metric, pixel_metric = self.make_category_data(category=category,)
            auroc_sp, f1_sp, ap_sp = image_metric
            auroc_px, f1_px, ap_px, aupro = pixel_metric
            auroc_sp_ls.append(auroc_sp)
            f1_sp_ls.append(f1_sp)
            ap_sp_ls.append(ap_sp)
            auroc_px_ls.append(auroc_px)
            f1_px_ls.append(f1_px)
            ap_px_ls.append(ap_px)
            aupro_ls.append(aupro)
        # mean
        auroc_sp_mean = sum(auroc_sp_ls) / len(auroc_sp_ls)
        f1_sp_mean = sum(f1_sp_ls) / len(f1_sp_ls)
        ap_sp_mean = sum(ap_sp_ls) / len(ap_sp_ls)
        auroc_px_mean = sum(auroc_px_ls) / len(auroc_px_ls)
        f1_px_mean = sum(f1_px_ls) / len(f1_px_ls)
        ap_px_mean = sum(ap_px_ls) / len(ap_px_ls)
        aupro_mean = sum(aupro_ls) / len(aupro_ls)

        for i, category in enumerate(self.categories):
            self.logger.info(category)
            self.logger.info('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp_ls[i]*100, f1_sp_ls[i]*100, ap_sp_ls[i]*100))
            self.logger.info('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px_ls[i]*100, f1_px_ls[i]*100, ap_px_ls[i]*100, aupro_ls[i]*100))
        self.logger.info('mean')
        self.logger.info('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp_mean*100, f1_sp_mean*100, ap_sp_mean*100))
        self.logger.info('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px_mean*100, f1_px_mean*100, ap_px_mean*100, aupro_mean*100))
        
        # save in excel
        if self.save_excel:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "MuSc_results"
            sheet.cell(row=1,column=2,value='auroc_px')
            sheet.cell(row=1,column=3,value='f1_px')
            sheet.cell(row=1,column=4,value='ap_px')
            sheet.cell(row=1,column=5,value='aupro')
            sheet.cell(row=1,column=6,value='auroc_sp')
            sheet.cell(row=1,column=7,value='f1_sp')
            sheet.cell(row=1,column=8,value='ap_sp')
            for col_index in range(2):
                for row_index in range(len(self.categories)):
                    if col_index == 0:
                        sheet.cell(row=row_index+2,column=col_index+1,value=self.categories[row_index])
                    else:
                        sheet.cell(row=row_index+2,column=col_index+1,value=auroc_px_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+2,value=f1_px_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+3,value=ap_px_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+4,value=aupro_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+5,value=auroc_sp_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+6,value=f1_sp_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+7,value=ap_sp_ls[row_index]*100)
                    if row_index == len(self.categories)-1:
                        if col_index == 0:
                            sheet.cell(row=row_index+3,column=col_index+1,value='mean')
                        else:
                            sheet.cell(row=row_index+3,column=col_index+1,value=auroc_px_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+2,value=f1_px_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+3,value=ap_px_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+4,value=aupro_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+5,value=auroc_sp_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+6,value=f1_sp_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+7,value=ap_sp_mean*100)
            workbook.save(os.path.join(self.output_dir, 'results.xlsx'))
class FeatureExtractor(torch.nn.Module):
    def __init__(self, backbone, layers_to_extract_from, device, input_shape,
                 cnn_backbones_most_dim, cnn_backbones_most_side):
        super(FeatureExtractor, self).__init__()
        self.backbone = backbone.to(device)
        self.layers_to_extract_from = layers_to_extract_from
        self.device = device
        self.input_shape = input_shape
        self.cnn_backbones_most_dim = cnn_backbones_most_dim
        self.cnn_backbones_most_side = cnn_backbones_most_side
        self.patch_maker = PatchMaker(3, stride=1)
        self.forward_modules = torch.nn.ModuleDict({})

        feature_aggregator = NetworkFeatureAggregator(
            self.backbone, self.layers_to_extract_from, self.device
        )
        feature_dimensions = feature_aggregator.feature_dimensions(input_shape)
        self.forward_modules["feature_aggregator"] = feature_aggregator

        preprocessing = Preprocessing(feature_dimensions, self.cnn_backbones_most_dim)
        self.forward_modules["preprocessing"] = preprocessing

        preadapt_aggregator = Aggregator(target_dim=self.cnn_backbones_most_dim)

        _ = preadapt_aggregator.to(self.device)

        self.forward_modules["preadapt_aggregator"] = preadapt_aggregator

        self.forward_modules.eval()

    @torch.no_grad()
    def embed_no_aggregation(self, images):
        """Returns feature embeddings for images."""

        _ = self.forward_modules["feature_aggregator"].eval()
        # feature extractor only backbone timing 
        # if torch.cuda.is_available():
        #     torch.cuda.synchronize()
        start_time = time.time()
        features = self.forward_modules["feature_aggregator"](images)
        # if torch.cuda.is_available():
        #     torch.cuda.synchronize()
        cnn_extract_time = time.time() - start_time
        
        features = [features[layer] for layer in self.layers_to_extract_from] # torch.Size([16, 448, 64, 64]) torch.Size([16, 1232, 32, 32])
        # self.logger.info('features.shape', features[0].shape)
        # self.logger.info('features.shape', features[1].shape)
        # self.logger.info('features.shape', features[2].shape)
        # self.logger.info('features.shape', features[3].shape)
        # Return raw multi-layer feature maps directly (skip patchify / alignment / aggregation).
        return features, cnn_extract_time
        
    @torch.no_grad()
    def embed(self, images):
        """Returns feature embeddings for images."""

        _ = self.forward_modules["feature_aggregator"].eval()
        # feature extractor only backbone timing 
        # if torch.cuda.is_available():
        #     torch.cuda.synchronize()
        start_time = time.time()
        features = self.forward_modules["feature_aggregator"](images)
        # if torch.cuda.is_available():
        #     torch.cuda.synchronize()
        cnn_extract_time = time.time() - start_time
        
        features = [features[layer] for layer in self.layers_to_extract_from] # torch.Size([16, 448, 64, 64]) torch.Size([16, 1232, 32, 32])
        # self.logger.info('features.shape', features[0].shape)
        # self.logger.info('features.shape', features[1].shape)
        # self.logger.info('features.shape', features[2].shape)
        # self.logger.info('features.shape', features[3].shape)
        features = [
            self.patch_maker.patchify(x, return_spatial_info=True) for x in
            features
        ]
        patch_shapes = [x[1] for x in features]
        features = [x[0] for x in features]
        ref_num_patches = patch_shapes[0]

        for i in range(1, len(features)):
            _features = features[i]
            patch_dims = patch_shapes[i]

            _features = _features.reshape(
                _features.shape[0], patch_dims[0], patch_dims[1],
                *_features.shape[2:]
            )
            _features = _features.permute(0, -3, -2, -1, 1, 2)
            perm_base_shape = _features.shape
            _features = _features.reshape(-1, *_features.shape[-2:])
            _features = F.interpolate(
                _features.unsqueeze(1),
                size=(ref_num_patches[0], ref_num_patches[1]),
                mode="bilinear",
                align_corners=False,
            )
            _features = _features.squeeze(1)
            _features = _features.reshape(
                *perm_base_shape[:-2], ref_num_patches[0], ref_num_patches[1]
            )
            _features = _features.permute(0, -2, -1, 1, 2, 3)
            _features = _features.reshape(len(_features), -1,
                                          *_features.shape[-3:])
            features[i] = _features
        features = [x.reshape(-1, *x.shape[-3:]) for x in features]
        # As different feature backbones & patching provide differently
        # sized features, these are brought into the correct form here.
        features = self.forward_modules["preprocessing"](features)
        features = self.forward_modules["preadapt_aggregator"](features)
        features = torch.reshape(features, (-1, self.cnn_backbones_most_side, self.cnn_backbones_most_side, self.cnn_backbones_most_dim))       
        features = torch.permute(features, (0, 3, 1, 2))
        
        return features, cnn_extract_time


# Image handling classes.
class PatchMaker:
    def __init__(self, patchsize, stride=None):
        self.patchsize = patchsize
        self.stride = stride

    def patchify(self, features, return_spatial_info=False):
        """Convert a tensor into a tensor of respective patches.
        Args:
            x: [torch.Tensor, bs x c x w x h]
        Returns:
            x: [torch.Tensor, bs * w//stride * h//stride, c, patchsize,
            patchsize]
        """
        padding = int((self.patchsize - 1) / 2)
        unfolder = torch.nn.Unfold(
            kernel_size=self.patchsize, stride=self.stride, padding=padding,
            dilation=1
        )
        unfolded_features = unfolder(features)
        number_of_total_patches = []
        for s in features.shape[-2:]:
            n_patches = (
                                s + 2 * padding - 1 * (self.patchsize - 1) - 1
                        ) / self.stride + 1
            number_of_total_patches.append(int(n_patches))
        unfolded_features = unfolded_features.reshape(
            *features.shape[:2], self.patchsize, self.patchsize, -1
        )
        unfolded_features = unfolded_features.permute(0, 4, 1, 2, 3)

        if return_spatial_info:
            return unfolded_features, number_of_total_patches
        return unfolded_features


class Preprocessing(torch.nn.Module):
    def __init__(self, input_dims, output_dim):
        super(Preprocessing, self).__init__()
        self.input_dims = input_dims
        self.output_dim = output_dim

        self.preprocessing_modules = torch.nn.ModuleList()
        for input_dim in input_dims:
            module = MeanMapper(output_dim)
            self.preprocessing_modules.append(module)

    def forward(self, features):
        _features = []
        for module, feature in zip(self.preprocessing_modules, features):
            _features.append(module(feature))
        return torch.stack(_features, dim=1)


class MeanMapper(torch.nn.Module):
    def __init__(self, preprocessing_dim):
        super(MeanMapper, self).__init__()
        self.preprocessing_dim = preprocessing_dim

    def forward(self, features):
        features = features.reshape(len(features), 1, -1)
        return F.adaptive_avg_pool1d(features,
                                     self.preprocessing_dim).squeeze(1)


class Aggregator(torch.nn.Module):
    def __init__(self, target_dim):
        super(Aggregator, self).__init__()
        self.target_dim = target_dim

    def forward(self, features):
        """Returns reshaped and average pooled features."""
        # batchsize x number_of_layers x input_dim -> batchsize x target_dim
        features = features.reshape(len(features), 1, -1)
        features = F.adaptive_avg_pool1d(features, self.target_dim)
        return features.reshape(len(features), -1)


class NetworkFeatureAggregator(torch.nn.Module):
    """Efficient extraction of network features."""

    def __init__(self, backbone, layers_to_extract_from, device):
        super(NetworkFeatureAggregator, self).__init__()
        """Extraction of network features.

        Runs a network only to the last layer of the list of layers where
        network features should be extracted from.

        Args:
            backbone: torchvision.model
            layers_to_extract_from: [list of str]
        """
        self.layers_to_extract_from = layers_to_extract_from
        self.backbone = backbone
        self.device = device
        if not hasattr(backbone, "hook_handles"):
            self.backbone.hook_handles = []
        for handle in self.backbone.hook_handles:
            handle.remove()
        self.outputs = {}

        for extract_layer in layers_to_extract_from:
            forward_hook = ForwardHook(
                self.outputs, extract_layer, layers_to_extract_from[-1]
            )
            if "." in extract_layer:
                extract_block, extract_idx = extract_layer.split(".")
                network_layer = backbone.__dict__["_modules"][extract_block]
                if extract_idx.isnumeric():
                    extract_idx = int(extract_idx)
                    network_layer = network_layer[extract_idx]
                else:
                    network_layer = network_layer.__dict__["_modules"][
                        extract_idx]
            else:
                network_layer = backbone.__dict__["_modules"][extract_layer]

            if isinstance(network_layer, torch.nn.Sequential):
                # Some models may expose empty Sequential containers (e.g. LSNet-T blocks1 when depth[0]=0).
                # In that case, register the hook on the container itself.
                if len(network_layer) > 0:
                    target = network_layer[-1]
                else:
                    target = network_layer
                self.backbone.hook_handles.append(
                    target.register_forward_hook(forward_hook)
                )
            else:
                self.backbone.hook_handles.append(
                    network_layer.register_forward_hook(forward_hook)
                )
        self.to(self.device)

    def forward(self, images):
        self.outputs.clear()
        with torch.no_grad():
            # The backbone will throw an Exception once it reached the last
            # layer to compute features from. Computation will stop there.
            try:
                _ = self.backbone(images)
            except LastLayerToExtractReachedException:
                pass
        return self.outputs

    def feature_dimensions(self, input_shape):
        """Computes the feature dimensions for all layers given input_shape."""
        _input = torch.ones([1] + list(input_shape)).to(self.device)
        _output = self(_input)
        return [_output[layer].shape[1] for layer in
                self.layers_to_extract_from]
    
class ForwardHook:
    def __init__(self, hook_dict, layer_name: str, last_layer_to_extract: str):
        self.hook_dict = hook_dict
        self.layer_name = layer_name
        self.raise_exception_to_break = copy.deepcopy(
            layer_name == last_layer_to_extract
        )

    def __call__(self, module, input, output):
        self.hook_dict[self.layer_name] = output
        if self.raise_exception_to_break:
            raise LastLayerToExtractReachedException()
        return None
class LastLayerToExtractReachedException(Exception):
    pass



